from openpyxl import load_workbook, Workbook
from collections import defaultdict


def CheckNone(e):
    if e is None:
        return ""
    else:
        return e.encode('utf-8')

def get_cell(x, y, sheet):
    return sheet.cell(row=x, column=y).value


def table_origin(column, column_check, data, row_start, row_end, sheet):
    for i in range(row_start, row_end):
        if get_cell(i, column_check, sheet) == "CHECK_L":
            category = get_cell(i, column, sheet)
        else:
            cell_name = get_cell(i, column, sheet)
            cell_unit = get_cell(i, column + 1, sheet)
            cell_mass = get_cell(i, column + 3, sheet)
            cell_material = get_cell(i, column + 4, sheet)
            cell_labor = get_cell(i, column + 6, sheet)
            cell_machine = get_cell(i, column + 7, sheet)
            data[category].append([cell_name, cell_unit, cell_mass,
                                   cell_material, cell_labor, cell_machine
                                   ])
    return data



def prepare_data(data, column_table_sub, column_check, row_start, row_end, sheet):
    for i in range(row_start, row_end):
        if get_cell(i, column_check, sheet) == "CHECK_R" and i == row_start:
            category = get_cell(i, column_table_sub, sheet)
            count = 0
        elif get_cell(i, column_check, sheet) == "CHECK_R" and i != row_start:
            data[category].append(count)
            category = get_cell(i, column_table_sub, sheet)
            count = 0
        else:
            cell_name = get_cell(i, column_table_sub, sheet)
            if cell_name != None:
                check_exist = False
                for m in data[category]:
                    if m[0] == cell_name:
                        check_exist = True
                if not check_exist:
                    cell_unit = get_cell(i, column_table_sub + 1, sheet)
                    cell_mass = get_cell(i, column_table_sub + 2, sheet)
                    cell_material = get_cell(i, column_table_sub + 3, sheet)
                    cell_labor = get_cell(i, column_table_sub + 4, sheet)
                    cell_machine = get_cell(i, column_table_sub + 5, sheet)
                    data[category].append([cell_name, cell_unit, cell_mass,
                                           cell_material, cell_labor, cell_machine
                                           ])
                    count += 1

    return data



if __name__ == '__main__':
    data = defaultdict(list)
    wb = load_workbook(filename='templates/03 ONG NUOC VA VAN 3384 x - rev 0.xlsx')
    column_table_origin = 5
    column_table_substitue = 70
    column_check_l = 68
    column_check_r = 69
    row_start_l = 7
    row_start_r = 7
    row_end = 754
    row_end_r = 754
    current_sheet = wb['So sanh']
    book = Workbook()
    sheet = book.active
    table = table_origin(column_table_origin, column_check_l, data,row_start_l,row_end,current_sheet)
    after_prepare = prepare_data(table,column_table_substitue, column_check_r, row_start_r, row_end_r, current_sheet)
    f = open("templates/text.txt","a+")
    after_prepare1 = sorted(after_prepare.keys())
    for category,sub in after_prepare.items():

        if type(after_prepare[category][-1]) != list:

            f.write('Dau Muc ' + category.encode('utf8')+ '\n')
            f.write("Change " + str(after_prepare[category][-1])+ '\n')
            for i in range(len(sub) - (1+after_prepare[category][-1])):
                if sub[i][0] is None:
                    f.write("None"+'\n')
                else:
                    f.write(CheckNone(sub[i][0]) + "---" + CheckNone(sub[i][1]) + "---" +
                        str(sub[i][2]) + "---" + str(sub[i][3]) + "---" + str(sub[i][4]) +
                            "---" + str(sub[i][5]) + '\n')
            f.write("Change Element" + '\n')
            for i in range(after_prepare[category][-1]):
                if sub[-(i+2)][0] is None:
                    f.write("None" + '\n')
                    # print("None")
                else:
                    f.write(CheckNone(sub[-(i+2)][0]) + "---" + CheckNone(sub[-(i+2)][1])+ "---" + str(sub[-(i+2)][2])+ "---" +
                            str(sub[-(i+2)][3]) + "---" + str(sub[-(i+2)][4]) + "---" + str(sub[-(i+2)][5]) + '\n')

        else:
            print(category)
            f.write('Dau Muc ' + category.encode('utf8') + '\n')
            f.write("Change 0 \n")
            for i in range(len(sub)):
                if sub[i][0] is None:
                    f.write("None" + '\n')
                    # write_cell(column_table_origin,row_end,"",current_sheet)
                else:
                    f.write(CheckNone(sub[i][0])+ "---"+ CheckNone(sub[i][1])+"---"+str(sub[i][2])+"---"+
                          str(sub[i][3])+"---"+ str(sub[i][4])+"---"+ str(sub[i][5]) + '\n')
                    # write_cell(column_table_origin, row_end+i, sub[i][0], current_sheet)
                    # write_cell(column_table_origin+1, row_end+i, sub[i][1], current_sheet)
                    # write_cell(column_table_origin+2, row_end+i, sub[i][3], current_sheet)
                    # write_cell(column_table_origin+3, row_end+i, sub[i][4], current_sheet)
                    # write_cell(column_table_origin+4, row_end+i, sub[i][5], current_sheet)

    f.close()



